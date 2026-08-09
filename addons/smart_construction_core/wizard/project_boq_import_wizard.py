# -*- coding: utf-8 -*-
import base64
import csv
import hashlib
import io
import logging
import re

from odoo import fields, models
from odoo.exceptions import UserError
from odoo.tools import misc

from ..models.support.state_guard import raise_guard

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


_logger = logging.getLogger(__name__)


class ProjectBoqImportWizard(models.TransientModel):
    _name = "project.boq.import.wizard"
    _description = "工程量清单导入"
    _rec_name = "filename"

    BATCH_CREATE_SIZE = 500

    UOM_ALIAS_MAP = {
        "m2": ["㎡", "m²", "平米", "平方米", "平方"],
        "m3": ["立方", "立方米"],
        "项": ["项", "项（包干）", "项(包干)", "item", "unit", "units"],
    }

    project_id = fields.Many2one(
        "project.project",
        string="项目",
        required=True,
    )
    section_type = fields.Selection(
        [
            ("building", "建筑"),
            ("installation", "安装/机电"),
            ("decoration", "装饰"),
            ("landscape", "景观"),
            ("municipal", "市政"),
            ("other", "其他"),
        ],
        string="工程类别",
        help="若文件未识别到工程类别，使用此处的默认值。",
    )
    boq_category = fields.Selection(
        [
            ("boq", "分部分项清单"),
            ("unit_measure", "单价措施清单"),
            ("total_measure", "总价措施清单"),
            ("fee", "规费"),
            ("tax", "税金"),
            ("other", "其他费用"),
        ],
        string="清单类别",
        default="boq",
        required=True,
        help="标识清单来源类别，避免分部分项与措施清单互相混淆。",
    )
    single_name = fields.Char("单项工程")
    unit_name = fields.Char("单位工程")
    source_type = fields.Selection(
        [
            ("tender", "招标清单"),
            ("contract", "合同清单"),
            ("settlement", "结算清单"),
        ],
        string="清单来源",
        default="contract",
    )
    version_code = fields.Char(
        "版本号",
        default=lambda self: "V1-%s" % fields.Date.context_today(self).strftime("%Y%m%d"),
        required=True,
    )
    state = fields.Selection(
        [("upload", "上传"), ("preview", "预检"), ("done", "完成")],
        default="upload",
        required=True,
        readonly=True,
    )
    preview_digest = fields.Char("文件摘要", readonly=True)
    preview_row_count = fields.Integer("解析行数", readonly=True)
    preview_item_count = fields.Integer("清单项数", readonly=True)
    preview_summary_count = fields.Integer("保留汇总行", readonly=True)
    preview_heading_count = fields.Integer("保留结构标题行", readonly=True)
    preview_calculation_detail_count = fields.Integer("保留费用计算明细", readonly=True)
    preview_skipped_count = fields.Integer(
        "忽略空白/辅助行",
        readonly=True,
        help="仅统计空白或无业务含义的辅助行；结构标题、页内小计和合计均完整保留。",
    )
    preview_warning_count = fields.Integer("警告数", readonly=True)
    preview_analysis_count = fields.Integer("综合单价分析数", readonly=True)
    preview_norm_line_count = fields.Integer("定额组成行数", readonly=True)
    preview_resource_line_count = fields.Integer("资源消耗行数", readonly=True)
    preview_summary_component_count = fields.Integer("单位工程汇总行数", readonly=True)
    parser_warning_log = fields.Text("源文件结构诊断", readonly=True)
    preview_amount = fields.Monetary(
        "预检合价", currency_field="currency_id", readonly=True
    )
    currency_id = fields.Many2one(
        "res.currency", related="project_id.company_id.currency_id", readonly=True
    )
    preview_log = fields.Text("预检结果", readonly=True)
    file = fields.Binary(
        string="导入文件",
        required=True,
        help="支持 XLS、XLSX、CSV；同一编码可形成多条清单行，导入生成独立草稿版本且不覆盖已发布清单。",
    )
    filename = fields.Char("文件名")
    log = fields.Text("导入日志", readonly=True)
    note = fields.Html(
        string="导入说明",
        readonly=True,
        default=lambda self: (
            "<ul>"
            "<li>同一编码在表中多次出现，将导入为多条清单行，并在工程结构中归入同一清单子目节点。</li>"
            "<li>若单位不存在，将自动规范化并创建新的计量单位。</li>"
            "<li>导入按文件摘要形成独立草稿版本，不覆盖任何已发布清单。</li>"
            "</ul>"
        ),
    )

    # -------------------------------------------------------------------------
    # 主入口
    # -------------------------------------------------------------------------
    def action_preflight(self):
        """Parse without business writes and freeze an exact file digest."""
        self.ensure_one()
        if not self.file:
            raise UserError("请先上传导入文件。")
        rows, pending_uoms, skipped, detail_payload = self.with_context(
            boq_import_preflight=True
        )._parse_file(include_details=True)
        if not rows:
            raise UserError(
                "未找到可导入的清单数据：请确认文件包含清单名称以及工程量、单价或金额。"
            )
        all_item_rows = [row for row in rows if (row.get("line_type") or "item") == "item"]
        calculation_detail_rows = [row for row in all_item_rows if row.get("is_calculation_detail")]
        item_rows = [row for row in all_item_rows if not row.get("is_calculation_detail")]
        summary_rows = [
            row for row in rows if row.get("source_row_type") in ("subtotal", "total")
        ]
        heading_rows = [row for row in rows if row.get("source_row_type") == "heading"]
        warnings = []
        missing_code = sum(1 for row in item_rows if not row.get("code"))
        missing_uom = sum(1 for row in item_rows if not row.get("uom_id"))
        if missing_code:
            warnings.append(f"{missing_code} 条清单项缺少编码")
        if missing_uom:
            warnings.append(f"{missing_uom} 条清单项的单位将在确认导入时创建或使用通用单位")
        if pending_uoms:
            warnings.append("待创建计量单位：" + "、".join(sorted(pending_uoms)))
        if self.parser_warning_log:
            warnings.append(
                "源 XLS 容器存在兼容性提示，数据已稳定解析；建议归档前另存为标准 XLSX。"
            )
        raw = base64.b64decode(self.file)
        amount = sum(
            (
                float(row.get("imported_amount") or 0.0)
                if row.get("has_imported_amount")
                else float(row.get("quantity") or 0.0) * float(row.get("price") or 0.0)
            )
            for row in item_rows
        )
        detected_single = next((row.get("single_name") for row in rows if row.get("single_name")), False)
        detected_unit = next((row.get("unit_name") for row in rows if row.get("unit_name")), False)
        detected_section = next((row.get("section_type") for row in rows if row.get("section_type")), False)
        self.write(
            {
                "state": "preview",
                "preview_digest": hashlib.sha256(raw).hexdigest(),
                "preview_row_count": len(rows),
                "preview_item_count": len(item_rows),
                "preview_summary_count": len(summary_rows),
                "preview_heading_count": len(heading_rows),
                "preview_calculation_detail_count": len(calculation_detail_rows),
                "preview_skipped_count": skipped,
                "preview_warning_count": len(warnings),
                "preview_analysis_count": len(detail_payload["analyses"]),
                "preview_norm_line_count": sum(
                    len(item["norm_lines"]) for item in detail_payload["analyses"]
                ),
                "preview_resource_line_count": sum(
                    len(item["resource_lines"]) for item in detail_payload["analyses"]
                ),
                "preview_summary_component_count": len(detail_payload["summary_components"]),
                "preview_amount": amount,
                **({"single_name": detected_single} if detected_single and not self.single_name else {}),
                **({"unit_name": detected_unit} if detected_unit and not self.unit_name else {}),
                **({"section_type": detected_section} if detected_section and not self.section_type else {}),
                "preview_log": "\n".join(
                    [
                        (
                            f"识别并保留 {len(rows)} 行，其中清单计价项 {len(item_rows)} 行、"
                            f"费用计算明细 {len(calculation_detail_rows)} 行、结构标题 {len(heading_rows)} 行、"
                            f"页内小计/合计 {len(summary_rows)} 行。"
                        ),
                        f"另忽略 {skipped} 行空白或无业务含义的辅助行。",
                        f"预检合价 {amount:.2f}。",
                        (
                            "识别综合单价分析 %s 项、定额组成 %s 行、资源消耗 %s 行。"
                            % (
                                len(detail_payload["analyses"]),
                                sum(len(item["norm_lines"]) for item in detail_payload["analyses"]),
                                sum(len(item["resource_lines"]) for item in detail_payload["analyses"]),
                            )
                        ),
                        "识别单位工程造价汇总 %s 行（与清单计价项分开保存）。"
                        % len(detail_payload["summary_components"]),
                        *(warnings or ["未发现阻断性问题。"]),
                    ]
                ),
            }
        )
        return {
            "type": "ir.actions.act_window",
            "name": "工程量清单导入",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError("请先上传导入文件。")
        if self.state != "preview" or not self.preview_digest:
            raise UserError("请先执行预检，再确认导入。")
        raw = base64.b64decode(self.file)
        digest = hashlib.sha256(raw).hexdigest()
        if digest != self.preview_digest:
            raise UserError("文件已发生变化，请重新执行预检。")
        if self.project_id and self.project_id.is_boq_frozen():
            raise_guard(
                "P0_BOQ_FROZEN",
                f"项目[{self.project_id.display_name}]",
                "导入新 BOQ 版本",
                reasons=["项目已进入结算/支付关键节点"],
                hints=["请先完成/撤销结算或付款流程后再导入新版本"],
            )

        rows, created_uoms, skipped, detail_payload = self._parse_file(include_details=True)
        if not rows:
            raise UserError(
                "未找到可导入的清单数据：\n"
                "请确认文件中至少有一行同时包含名称列（清单名称/项目名称/汇总内容等）"
                "并且数量/单价/金额至少有一个为数字。"
            )

        Version = self.env["project.boq.version"]
        version_code = (self.version_code or "").strip()
        if not version_code:
            raise UserError("请填写版本号。")
        if Version.search_count(
            [
                ("project_id", "=", self.project_id.id),
                ("source_type", "=", self.source_type),
                ("code", "=", version_code),
            ]
        ):
            raise UserError("同一项目和清单来源下已存在该版本号，请使用新的版本号。")
        version = Version.create(self._prepare_version_values(version_code))
        batch = self.env["project.boq.import.batch"].create(
            {
                "name": f"{version_code} · {self.filename or '清单导入'}",
                "project_id": self.project_id.id,
                "version_id": version.id,
                "filename": self.filename or "未命名文件",
                "file_digest": digest,
                "row_count": len(rows),
                "item_count": self.preview_item_count,
                "skipped_count": skipped,
                "warning_count": self.preview_warning_count,
                "preview_payload": {
                    "schema": "sc.boq.import.preview.v1",
                    "row_count": len(rows),
                    "item_count": self.preview_item_count,
                    "summary_count": self.preview_summary_count,
                    "heading_count": self.preview_heading_count,
                    "calculation_detail_count": self.preview_calculation_detail_count,
                    "skipped_count": skipped,
                    "warning_count": self.preview_warning_count,
                    "amount": self.preview_amount,
                    "source_diagnostics": (self.parser_warning_log or "").splitlines(),
                    "analysis_count": len(detail_payload["analyses"]),
                    "norm_line_count": sum(
                        len(item["norm_lines"]) for item in detail_payload["analyses"]
                    ),
                    "resource_line_count": sum(
                        len(item["resource_lines"]) for item in detail_payload["analyses"]
                    ),
                    "summary_component_count": len(detail_payload["summary_components"]),
                },
            }
        )
        for vals in rows:
            vals.update({"version_id": version.id, "import_batch_id": batch.id})

        Boq = self.env["project.boq.line"]

        def _create_rows(vals_list):
            """按行的 boq_category 决定是否启用层级导入。"""
            if not vals_list:
                return 0

            grouped = {}
            for vals in vals_list:
                cat = vals.get("boq_category") or self.boq_category or "boq"
                grouped.setdefault(cat, []).append(vals)

            created = 0
            for cat, chunk in grouped.items():
                if cat in ("boq", "other"):
                    created += self._create_with_hierarchy(Boq, chunk)
                else:
                    created += self._batch_create(Boq, chunk)
            return created

        created_count = _create_rows(rows)
        analysis_count = self._create_analysis_snapshot(version, detail_payload)
        summary_component_count = self._create_summary_snapshot(version, detail_payload)
        self._finalize_source_summary_calculations(version.line_ids)

        log_lines = []
        log_lines.append(f"成功导入 {created_count} 条到清单版本 {version_code}。")
        if analysis_count:
            log_lines.append(
                "关联综合单价分析 %s 项、定额组成 %s 行、资源消耗 %s 行。"
                % (
                    analysis_count,
                    sum(len(item["norm_lines"]) for item in detail_payload["analyses"]),
                    sum(len(item["resource_lines"]) for item in detail_payload["analyses"]),
                )
            )
        if summary_component_count:
            log_lines.append("保留单位工程造价汇总 %s 行。" % summary_component_count)
        if skipped:
            log_lines.append(f"跳过 {skipped} 行（空行/小计行/无数值行）。")
        if created_uoms:
            log_lines.append("自动创建计量单位：\n- " + "\n- ".join(sorted(created_uoms)))
        log_lines.append("版本已校验；发布后可在 WBS 计划中按管理目标分配清单来源。")
        self.log = "\n".join(log_lines)
        batch.write(
            {
                "state": "imported",
                "log": self.log,
                "imported_at": fields.Datetime.now(),
                "imported_by": self.env.user.id,
            }
        )
        version.action_validate()
        self.state = "done"

        return {
            "type": "ir.actions.act_window",
            "res_model": "project.boq.version",
            "res_id": version.id,
            "view_mode": "form",
            "target": "current",
        }

    # -------------------------------------------------------------------------
    # 文件解析入口
    # -------------------------------------------------------------------------
    def _parse_file(self, include_details=False):
        """Parse CSV/XLS/XLSX into vals list for project.boq.line."""
        data = base64.b64decode(self.file)
        filename = (self.filename or "").lower()
        parser = BoqParser(self)
        return parser.parse_file(data, filename, include_details=include_details)

    def _prepare_version_values(self, version_code):
        """Stable extension hook for optional industry-support modules."""
        return {
            "name": f"{self.project_id.display_name} {version_code}",
            "code": version_code,
            "project_id": self.project_id.id,
            "source_type": self.source_type,
        }

    @staticmethod
    def _analysis_match_key(values):
        def norm(value):
            return re.sub(r"\s+", "", str(value or "")).lower()

        return (
            norm(values.get("single_name")),
            norm(values.get("unit_name")),
            norm(values.get("major_name")),
            norm(values.get("boq_code") or values.get("source_code") or values.get("code")),
        )

    def _create_analysis_snapshot(self, version, detail_payload):
        analyses = detail_payload.get("analyses") or []
        if not analyses:
            return 0
        line_pool = {}
        item_lines = version.line_ids.filtered(lambda line: line.line_type == "item")
        for line in item_lines.sorted(lambda row: (row.sheet_index or 0, row.sequence or 0, row.id)):
            key = self._analysis_match_key(
                {
                    "single_name": line.single_name,
                    "unit_name": line.unit_name,
                    "major_name": line.major_name,
                    "boq_code": line.source_code or line.code,
                }
            )
            line_pool.setdefault(key, []).append(line)

        Analysis = self.env["project.boq.analysis"]
        created = Analysis.browse()
        unmatched = []
        for payload in analyses:
            candidates = line_pool.get(self._analysis_match_key(payload), [])
            if not candidates:
                unmatched.append(
                    "%s / %s / %s"
                    % (payload.get("unit_name") or "-", payload.get("major_name") or "-", payload["boq_code"])
                )
                continue
            boq_line = candidates.pop(0)
            created |= Analysis.create(
                {
                    "name": payload.get("name") or boq_line.name,
                    "boq_line_id": boq_line.id,
                    "uom_raw": payload.get("uom_raw"),
                    "source_quantity": payload.get("source_quantity", 0.0),
                    "source_unit_price": payload.get("source_unit_price", 0.0),
                    "source_sheet_index": payload.get("source_sheet_index"),
                    "source_sheet_name": payload.get("source_sheet_name"),
                    "source_sequence": payload.get("source_sequence"),
                    "single_name": payload.get("single_name"),
                    "unit_name": payload.get("unit_name"),
                    "major_name": payload.get("major_name"),
                    "norm_line_ids": [(0, 0, vals) for vals in payload.get("norm_lines", [])],
                    "resource_line_ids": [
                        (0, 0, vals) for vals in payload.get("resource_lines", [])
                    ],
                }
            )
        if unmatched:
            raise UserError(
                "有 %s 项综合单价分析无法与清单项关联，请核对单项工程、单位工程、专业和清单编码。\n%s"
                % (len(unmatched), "\n".join(unmatched[:20]))
            )
        created._resolve_norm_links()
        return len(created)

    def _create_summary_snapshot(self, version, detail_payload):
        values = detail_payload.get("summary_components") or []
        if not values:
            return 0
        self.env["project.boq.summary.component"].create(
            [dict(row, version_id=version.id) for row in values]
        )
        return len(values)

    def _parse_csv_content(self, content):
        reader = csv.reader(io.StringIO(content))
        rows_data = list(reader)
        if not rows_data:
            raise UserError("导入文件没有数据，请检查。")
        # 头部探测：使用第一行作为表头
        headers = [str(h or "").strip() for h in rows_data[0]]
        data_rows = rows_data[1:]
        col_map = self._prepare_col_map(headers)
        rows, created_uoms, skipped = self._build_rows_from_iter(
            data_rows,
            col_map,
            strict_numeric=True,
            default_single=self.single_name,
            default_unit=self.unit_name,
            boq_category=self.boq_category,
        )
        if not rows:
            rows, created_uoms, skipped = self._build_rows_from_iter(
                data_rows,
                col_map,
                strict_numeric=False,
                default_single=self.single_name,
                default_unit=self.unit_name,
                boq_category=self.boq_category,
            )
        return rows, created_uoms, skipped

    def _parse_excel(self, data, filename, include_details=False):
        """解析 XLS/XLSX 为 project.boq.line 的 vals 列表。"""
        col_map_cfg = self._col_map_cfg()
        rows_all = []
        created_uoms_all = set()
        skipped_all = 0
        detail_payload = {
            "schema": "sc.boq.full_import.v1",
            "analyses": [],
            "summary_components": [],
        }

        # ---------------- XLSX ----------------
        if filename.endswith(".xlsx"):
            if not openpyxl:
                raise UserError("服务器缺少 openpyxl，无法解析 XLSX，请安装依赖或改用 CSV。")
            book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

            for idx, sheet in enumerate(book.worksheets, start=1):
                title = sheet.title or ""
                # 根据 sheet 名称分类；封面/汇总等直接跳过
                sheet_type, sheet_category = self._classify_sheet_title(title)
                if sheet_type in ("cover", "summary", "other_skip"):
                    continue

                if not self._is_supported_sheet(title):
                    continue

                if sheet_type in ("boq_analysis", "measure_analysis"):
                    parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_excel(
                        sheet, limit=6
                    )
                    detail_payload["analyses"].extend(
                        self._parse_analysis_rows(
                            [list(row) for row in sheet.iter_rows(values_only=True)],
                            source_sheet_index=idx,
                            source_sheet_name=title,
                            single_name=self.single_name or parsed_single,
                            unit_name=self.unit_name or parsed_unit,
                            major_name=parsed_major,
                        )
                    )
                    continue

                if sheet_type == "unit_summary":
                    parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_excel(
                        sheet, limit=6
                    )
                    detail_payload["summary_components"].extend(
                        self._parse_unit_summary_rows(
                            [list(row) for row in sheet.iter_rows(values_only=True)],
                            source_sheet_index=idx,
                            source_sheet_name=title,
                            single_name=self.single_name or parsed_single,
                            unit_name=self.unit_name or parsed_unit,
                            major_name=parsed_major,
                        )
                    )
                    continue

                # 表头行数按表类型区分：总价/规费/税金等通常只有 1 行列标题
                if sheet_type in ("total_measure", "fee", "tax", "other_item"):
                    header_info = self._extract_excel_header(sheet, header_rows=1)
                else:
                    header_info = self._extract_excel_header(sheet)
                if not header_info:
                    continue
                headers, data_start_row, header_row_idx = header_info

                # 解析表头前几行的“工程名称：项目\单位【专业】”
                parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_excel(
                    sheet, limit=max(5, header_row_idx)
                )

                default_single = self.single_name or parsed_single
                default_unit = self.unit_name or parsed_unit

                section_type = (
                    self.section_type
                    or self._map_major_to_section_type(parsed_major)
                    or self._guess_section_type(title)
                )
                # 根据 sheet 名推断清单类别：分部分项 / 单价措施 / 总价措施 / 规费 / 税金 / 其他项目
                category = sheet_category or self._detect_boq_category(title) or self.boq_category

                col_map = self._prepare_col_map(headers, col_map_cfg)
                data_rows = [
                    list(row)
                    for row in sheet.iter_rows(min_row=data_start_row, values_only=True)
                ]

                # 清单类 sheet 统一 strict_numeric=False，
                # 标题/分部行会先保留，由内部的小计/合计过滤逻辑做最后筛选
                if category == "other":
                    rows, skipped = self._build_rows_other(
                        data_rows,
                        sheet_index=idx,
                        sheet_name=title,
                        section_type=section_type,
                        default_single=default_single,
                        default_unit=default_unit,
                        major_name=parsed_major,
                    )
                    rows_all.extend(rows)
                    skipped_all += skipped
                    continue

                rows, created_uoms, skipped = self._build_rows_from_iter(
                    data_rows,
                    col_map,
                    section_type=section_type,
                    strict_numeric=False,
                    default_single=default_single,
                    default_unit=default_unit,
                    major_name=parsed_major,
                    sheet_index=idx,
                    sheet_name=title,
                    boq_category=category,
                )

                rows_all.extend(rows)
                created_uoms_all.update(created_uoms)
                skipped_all += skipped

            result = (rows_all, created_uoms_all, skipped_all)
            return (*result, detail_payload) if include_details else result

        # ---------------- XLS ----------------
        if filename.endswith(".xls"):
            if not xlrd:
                raise UserError("服务器缺少 xlrd，无法解析 XLS，请安装依赖或改用 CSV。")
            diagnostic_stream = io.StringIO()
            book = xlrd.open_workbook(file_contents=data, logfile=diagnostic_stream)
            diagnostics = self._normalize_xls_diagnostics(diagnostic_stream.getvalue())
            self.parser_warning_log = "\n".join(diagnostics) if diagnostics else False

            for idx, sheet in enumerate(book.sheets(), start=1):
                if sheet.nrows < 1:
                    continue

                title = sheet.name or ""
                sheet_type, sheet_category = self._classify_sheet_title(title)
                if sheet_type in ("cover", "summary", "other_skip"):
                    continue

                if not self._is_supported_sheet(title):
                    continue

                if sheet_type in ("boq_analysis", "measure_analysis"):
                    parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_xls(
                        sheet, limit=6
                    )
                    detail_payload["analyses"].extend(
                        self._parse_analysis_rows(
                            [
                                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                                for r in range(sheet.nrows)
                            ],
                            source_sheet_index=idx,
                            source_sheet_name=title,
                            single_name=self.single_name or parsed_single,
                            unit_name=self.unit_name or parsed_unit,
                            major_name=parsed_major,
                        )
                    )
                    continue

                if sheet_type == "unit_summary":
                    parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_xls(
                        sheet, limit=6
                    )
                    detail_payload["summary_components"].extend(
                        self._parse_unit_summary_rows(
                            [
                                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                                for r in range(sheet.nrows)
                            ],
                            source_sheet_index=idx,
                            source_sheet_name=title,
                            single_name=self.single_name or parsed_single,
                            unit_name=self.unit_name or parsed_unit,
                            major_name=parsed_major,
                        )
                    )
                    continue

                if sheet_type in ("total_measure", "fee", "tax", "other_item"):
                    headers, data_start_row, header_row_idx = self._extract_xls_header(sheet, header_rows=1)
                else:
                    headers, data_start_row, header_row_idx = self._extract_xls_header(sheet)
                if not headers:
                    continue

                parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_xls(
                    sheet, limit=max(5, header_row_idx)
                )

                default_single = self.single_name or parsed_single
                default_unit = self.unit_name or parsed_unit

                section_type = (
                    self.section_type
                    or self._map_major_to_section_type(parsed_major)
                    or self._guess_section_type(title)
                )
                category = sheet_category or self._detect_boq_category(title) or self.boq_category

                col_map = self._prepare_col_map(headers, col_map_cfg)
                data_rows = [
                    [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    for r in range(data_start_row, sheet.nrows)
                ]

                if category == "other":
                    rows, skipped = self._build_rows_other(
                        data_rows,
                        sheet_index=idx,
                        sheet_name=title,
                        section_type=section_type,
                        default_single=default_single,
                        default_unit=default_unit,
                        major_name=parsed_major,
                    )
                    rows_all.extend(rows)
                    skipped_all += skipped
                    continue

                rows, created_uoms, skipped = self._build_rows_from_iter(
                    data_rows,
                    col_map,
                    section_type=section_type,
                    strict_numeric=False,
                    default_single=default_single,
                    default_unit=default_unit,
                    major_name=parsed_major,
                    sheet_index=idx,
                    sheet_name=title,
                    boq_category=category,
                )

                rows_all.extend(rows)
                created_uoms_all.update(created_uoms)
                skipped_all += skipped

            result = (rows_all, created_uoms_all, skipped_all)
            return (*result, detail_payload) if include_details else result

        # ---------------- Fallback: 按 CSV 解析 ----------------
        result = (self._parse_csv_bytes(data), set(), 0)
        return (*result, detail_payload) if include_details else result

    def _parse_analysis_rows(
        self,
        rows,
        source_sheet_index,
        source_sheet_name,
        single_name=None,
        unit_name=None,
        major_name=None,
    ):
        """Parse standard 表-09 blocks without leaking workbook layout into models."""
        analyses = []
        current = None
        mode = None

        def cell(row, index):
            return row[index] if index < len(row) else ""

        def text(value):
            return str(value or "").strip()

        def compact(value):
            return re.sub(r"\s+", "", text(value))

        def finish():
            nonlocal current
            if not current:
                return
            if not current["source_unit_price"]:
                current["source_unit_price"] = sum(
                    line["amount_labor"]
                    + line["amount_material"]
                    + line["amount_machine"]
                    + line["amount_overhead"]
                    + line["amount_profit"]
                    for line in current["norm_lines"]
                )
            analyses.append(current)
            current = None

        for row_number, row in enumerate(rows, start=1):
            first = compact(cell(row, 0))
            if first == "项目编码":
                finish()
                raw_code = text(cell(row, 1))
                matches = re.findall(r"\d{12}", raw_code)
                if not matches:
                    mode = None
                    continue
                current = {
                    "boq_code": matches[-1],
                    "name": text(cell(row, 5)) or text(cell(row, 3)),
                    "uom_raw": text(cell(row, 9)) or text(cell(row, 8)),
                    "source_quantity": self._to_float(cell(row, 13)),
                    "source_unit_price": 0.0,
                    "source_sheet_index": source_sheet_index,
                    "source_sheet_name": source_sheet_name,
                    "source_sequence": len(analyses) + 1,
                    "single_name": single_name or False,
                    "unit_name": unit_name or False,
                    "major_name": major_name or False,
                    "norm_lines": [],
                    "resource_lines": [],
                }
                mode = None
                continue
            if not current:
                continue
            if first == "定额编号":
                mode = "norm"
                continue
            if "材料费明细" in first:
                mode = "material"
                continue
            if "清单项目综合单价" in first:
                current["source_unit_price"] = self._to_float(cell(row, 9))
                mode = None
                continue
            if first in ("小计", "未计价材料费"):
                mode = None
                continue
            if mode == "norm":
                code = text(cell(row, 0))
                name = text(cell(row, 1))
                if code and name and self._is_number(cell(row, 3)):
                    current["norm_lines"].append(
                        {
                            "sequence": len(current["norm_lines"]) + 1,
                            "norm_code": code,
                            "name": name,
                            "unit_raw": text(cell(row, 2)),
                            "budget_consumption": self._to_float(cell(row, 3)),
                            "unit_labor": self._to_float(cell(row, 4)),
                            "unit_material": self._to_float(cell(row, 5)),
                            "unit_machine": self._to_float(cell(row, 6)),
                            "unit_overhead": self._to_float(cell(row, 7)),
                            "unit_profit": self._to_float(cell(row, 8)),
                            "amount_labor": self._to_float(cell(row, 9)),
                            "amount_material": self._to_float(cell(row, 10)),
                            "amount_machine": self._to_float(cell(row, 11)),
                            "amount_overhead": self._to_float(cell(row, 12)),
                            "amount_profit": self._to_float(cell(row, 13)),
                            "source_row": row_number,
                        }
                    )
                continue
            if mode == "material":
                name = text(cell(row, 1))
                if "材料费小计" in name:
                    mode = None
                    continue
                if not name or "主要材料名称" in name:
                    continue
                unit = text(cell(row, 7))
                consumption = self._to_float(cell(row, 8))
                price = self._to_float(cell(row, 9))
                source_amount = self._to_float(cell(row, 10))
                if not (unit or consumption or price or source_amount):
                    continue
                if not consumption and source_amount:
                    consumption = 1.0
                    price = source_amount
                current["resource_lines"].append(
                    {
                        "sequence": len(current["resource_lines"]) + 1,
                        "resource_type": "material",
                        "name": name,
                        "unit_raw": unit,
                        "budget_consumption": consumption,
                        "budget_unit_price": price,
                        "budget_unit_amount": source_amount,
                        "provisional_unit_price": self._to_float(cell(row, 11)),
                        "provisional_unit_amount": self._to_float(cell(row, 13)),
                        "source_row": row_number,
                    }
                )
        finish()
        return analyses

    def _parse_unit_summary_rows(
        self,
        rows,
        source_sheet_index,
        source_sheet_name,
        single_name=None,
        unit_name=None,
        major_name=None,
    ):
        """Keep E.3 unit-project summaries separate from BOQ leaves."""
        result = []
        for row_number, row in enumerate(rows, start=1):
            code = str((row[0] if len(row) > 0 else "") or "").strip()
            name = str((row[1] if len(row) > 1 else "") or "").strip()
            amount_raw = row[2] if len(row) > 2 else ""
            provisional_raw = row[3] if len(row) > 3 else ""
            if "总价合计" in code:
                name = code
                code = ""
                if not self._is_number(amount_raw):
                    amount_raw = row[1] if len(row) > 1 else ""
                    provisional_raw = row[2] if len(row) > 2 else ""
            if not name or not self._is_number(amount_raw):
                continue
            if code == "1":
                component_type = "direct"
            elif code == "2":
                component_type = "measure"
            elif code == "3":
                component_type = "other"
            elif code == "4":
                component_type = "fee"
            elif code == "6":
                component_type = "pre_tax"
            elif code in {"7", "8"}:
                component_type = "tax"
            elif not code and "总价合计" in name:
                component_type = "total"
            else:
                component_type = "detail"
            result.append(
                {
                    "sequence": len(result) + 1,
                    "code": code or False,
                    "name": name,
                    "component_type": component_type,
                    "amount": self._to_float(amount_raw),
                    "provisional_amount": self._to_float(provisional_raw),
                    "single_name": single_name or False,
                    "unit_name": unit_name or False,
                    "major_name": major_name or False,
                    "source_sheet_index": source_sheet_index,
                    "source_sheet_name": source_sheet_name,
                    "source_row": row_number,
                }
            )
        pre_tax = next(
            (row["amount"] for row in result if row["component_type"] == "pre_tax"),
            0.0,
        )
        if pre_tax:
            for row in result:
                if row["component_type"] == "tax":
                    row["calc_base"] = "税前不含税工程造价"
                    row["source_rate"] = row["amount"] / pre_tax * 100.0
        return result

    # -------------------------------------------------------------------------
    # Sheet 名称分类（核心升级点）
    # -------------------------------------------------------------------------
    @staticmethod
    def _classify_sheet_title(sheet_title):
        """
        根据 sheet 名称判断：
        - sheet_type: 业务上的表类型（boq/unit_measure/total_measure/fee/tax/other）
        - category:   写入 project.boq.line.boq_category 的值

        返回: (sheet_type, category)；都可能是 None。
        """
        title_raw = sheet_title or ""
        # 去掉空格/全角空格，全部小写方便匹配
        text = title_raw.replace(" ", "").replace("\u3000", "").lower()

        # 综合单价分析必须先于普通清单匹配，否则定额行会被污染为 BOQ 行。
        if "综合单价分析表" in text:
            if "措施项目" in text:
                return "measure_analysis", "unit_measure"
            return "boq_analysis", "boq"

        if "单位工程招标控制价投标报价汇总表" in text or "单位工程投标报价汇总表" in text:
            return "unit_summary", None

        # 1) 总价措施项目清单计价表（优先匹配，避免被“措施”二字误判成单价措施）
        if "总价措施项目清单计价表" in text or "总价措施项目清单" in text:
            return "total_measure", "total_measure"

        # 2) 单价措施项目清单
        if "单价措施项目清单" in text or "单价措施" in text:
            return "unit_measure", "unit_measure"

        # 3) 分部分项工程量清单
        if "分部分项工程量清单" in text or "分部分项工程清单" in text:
            return "boq", "boq"

        # 4) 其他项目清单 / 计价汇总表
        if "其他项目清单计价汇总表" in text or "其他项目清单" in text or "其他项目" in text:
            return "other_item", "other"

        # 5) 规费/税金 等专门表（有的模板会单独拆出来）
        if "规费" in text and "清单" in text:
            return "fee", "fee"
        if "税金" in text and "清单" in text:
            return "tax", "tax"

        # 6) 其它含“措施”但没说总价/单价的，尽量保守按 total_measure 处理
        if "总价措施" in text:
            return "total_measure", "total_measure"
        if "措施项目清单" in text:
            # 如果没提“单价/总价”，按单价措施兜底
            return "unit_measure", "unit_measure"

        # 7) 实在看不出来，一律当分部分项
        return None, None
    # -------------------------------------------------------------------------
    # 表头 & 列映射
    # -------------------------------------------------------------------------
    def _col_map_cfg(self):
        return {
            "code": ["清单编码", "编码", "code"],
            "name": ["清单名称", "名称", "name", "项目名称", "汇总内容"],
            "spec": ["规格", "规格型号", "项目特征", "项目特征描述"],
            "uom_id": ["单位", "uom"],
            "quantity": ["工程量", "数量", "qty"],
            "price": ["单价", "price"],
            # “金额（元）”多见于总标题，不直接当金额列匹配
            "amount": ["合价", "合计", "amount", "金额", "金额元"],
            "cost_item_id": ["成本项", "成本科目"],
            "remark": ["备注", "说明"],
            # --- 总价措施/规费类专用（目前先读出来，后面可扩展成模型字段） ---
            "rate": ["费率", "费率(%)", "费率（%）"],
            "calc_base": ["计算基础", "计费基础"],
        }

    def _prepare_col_map(self, headers, col_map_cfg=None):
        col_map_cfg = col_map_cfg or self._col_map_cfg()
        col_map = {}
        for idx, title in enumerate(headers):
            title_norm = self._normalize_header(title)
            for field, aliases in col_map_cfg.items():
                matched = False
                for alias in aliases:
                    alias_norm = self._normalize_header(alias)
                    if (
                        title_norm == alias_norm
                        or title_norm.endswith(alias_norm)
                        or alias_norm in title_norm
                    ):
                        matched = True
                        break
                if matched and field not in col_map:
                    col_map[field] = idx
                    break
        if "name" not in col_map:
            # 兜底：首列作为名称
            if headers:
                col_map["name"] = 0
            else:
                raise UserError("模板中至少需要包含 “清单名称” 列。")
        # 若识别到工程量列，按相对位置推断单价/合价（常见 F1-1 结构：工程量右一列=单价，右二列=合价）
        qty_idx = col_map.get("quantity")
        if qty_idx is not None:
            if "price" not in col_map and qty_idx + 1 < len(headers):
                col_map["price"] = qty_idx + 1
            if "amount" not in col_map and qty_idx + 2 < len(headers):
                col_map["amount"] = qty_idx + 2
        return col_map

    def _find_header_in_sheet(self, row_iter):
        """Deprecated: use _extract_excel_header/_extract_xls_header."""
        return [], 0

    # ----------------- Excel helpers -----------------
    def _parse_engineering_header_excel(self, sheet, limit=5):
        """
        解析表头中的“工程名称：项目\\单位【专业】”
        返回 (single_name, unit_name, major_name)
        """
        for row in sheet.iter_rows(min_row=1, max_row=limit, values_only=True):
            values = list(row)
            for index, val in enumerate(values):
                if not val:
                    continue
                text = str(val)
                if "工程名称" not in text:
                    continue
                payload = self._engineering_header_payload(text, values[index + 1 :])
                if payload:
                    return self._split_engineering_header(payload)
        return "", "", ""

    def _parse_engineering_header_xls(self, sheet, limit=5):
        for r in range(min(limit, sheet.nrows)):
            values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for index, val in enumerate(values):
                if not val:
                    continue
                text = str(val)
                if "工程名称" not in text:
                    continue
                payload = self._engineering_header_payload(text, values[index + 1 :])
                if payload:
                    return self._split_engineering_header(payload)
        return "", "", ""

    @staticmethod
    def _engineering_header_payload(label_cell, following_cells):
        payload = str(label_cell or "").split("工程名称", 1)[-1].lstrip("：:").strip()
        if payload:
            return payload
        return next((str(value).strip() for value in following_cells if str(value or "").strip()), "")

    @staticmethod
    def _split_engineering_header(payload):
        parts = str(payload or "").split("\\", 1)
        single = parts[0].strip() if parts else ""
        tail = parts[1].strip() if len(parts) >= 2 else ""
        unit = tail
        major = ""
        if "【" in tail and "】" in tail:
            before, suffix = tail.split("【", 1)
            unit = before.strip()
            major = suffix.split("】", 1)[0].strip()
        return single, unit, major

    def _extract_excel_header(self, sheet, header_rows=2, scan_rows=8):
        """处理多行表头+合并单元格，返回(扁平列名列表, 数据起始行号, 识别到的表头行号)"""
        merge_map = {}
        try:
            merge_ranges = sheet.merged_cells
        except Exception:
            merge_ranges = None
        if merge_ranges:
            ranges = getattr(merge_ranges, "ranges", merge_ranges)
            try:
                for m in ranges:
                    min_row, min_col, max_row, max_col = m.min_row, m.min_col, m.max_row, m.max_col
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            merge_map[(r, c)] = (min_row, min_col)
            except Exception:
                merge_map = {}

        def cell_val(r, c):
            key = merge_map.get((r, c))
            if key:
                r, c = key
            return sheet.cell(row=r, column=c).value

        max_col = sheet.max_column or 0
        header_row_idx = 0
        best_hits = 0
        keywords = [
            "编码",
            "项目编码",
            "清单编码",
            "特征",
            "工程量",
            "综合单价",
            "合价",
            "计量单位",
            # 其他项目清单/计价汇总表常见列
            "项目名称",
            "金额",
            "金额(元)",
            "金额（元）",
            "备注",
            "序号",
        ]
        for idx in range(1, min(scan_rows, sheet.max_row or 0) + 1):
            row_vals = [str(cell_val(idx, c) or "").strip() for c in range(1, max_col + 1)]
            hits = sum(1 for v in row_vals if any(k in v for k in keywords))
            if hits > best_hits:
                best_hits = hits
                header_row_idx = idx
        if not header_row_idx:
            return None

        header_rows_vals = []
        for r in range(header_row_idx, min(header_row_idx + header_rows, (sheet.max_row or 0) + 1)):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(str(cell_val(r, c) or "").strip())
            header_rows_vals.append(row_vals)

        # 纵向拼接列名
        flat_headers = []
        for c in range(max_col):
            parts = []
            for r in range(len(header_rows_vals)):
                v = header_rows_vals[r][c]
                if v:
                    parts.append(v)
            flat_headers.append(" - ".join(parts) if parts else "")

        data_start = header_row_idx + header_rows
        return flat_headers, data_start, header_row_idx

    def _extract_xls_header(self, sheet, header_rows=2, scan_rows=8):
        max_col = sheet.ncols
        header_row_idx = 0
        best_hits = 0
        keywords = [
            "编码",
            "项目编码",
            "清单编码",
            "特征",
            "工程量",
            "综合单价",
            "合价",
            "计量单位",
            # 其他项目清单/计价汇总表常见列
            "项目名称",
            "金额",
            "金额(元)",
            "金额（元）",
            "备注",
            "序号",
        ]
        for idx in range(min(scan_rows, sheet.nrows)):
            row_vals = [str(sheet.cell_value(idx, c) or "").strip() for c in range(max_col)]
            hits = sum(1 for v in row_vals if any(k in v for k in keywords))
            if hits > best_hits:
                best_hits = hits
                header_row_idx = idx
        if max_col == 0:
            return None, 0, 0
        header_rows_vals = []
        for r in range(header_row_idx, min(header_row_idx + header_rows, sheet.nrows)):
            row_vals = [str(sheet.cell_value(r, c) or "").strip() for c in range(max_col)]
            header_rows_vals.append(row_vals)
        flat_headers = []
        for c in range(max_col):
            parts = [row_vals[c] for row_vals in header_rows_vals if row_vals[c]]
            flat_headers.append(" - ".join(parts) if parts else "")
        data_start = header_row_idx + header_rows
        return flat_headers, data_start, header_row_idx

    # -------------------------------------------------------------------------
    # 行解析 & 单位/成本项匹配
    # -------------------------------------------------------------------------
    def _build_rows_from_iter(
        self,
        row_iter,
        col_map,
        section_type=None,
        strict_numeric=True,
        default_single=None,
        default_unit=None,
        major_name=None,
        sheet_index=None,
        sheet_name=None,
        boq_category=None,
    ):
        # Unit creation is an import-service responsibility.  Elevate only
        # this bounded master-data lookup/create path instead of granting cost
        # users Inventory Administrator rights.
        Uom = self.env["uom.uom"].sudo()
        Dict, dict_domain_key = self._get_dictionary_model()

        rows = []
        uom_cache = {}
        cost_item_cache = {}
        created_uoms = set()
        skipped_rows = 0
        current_division = None

        def _default_uom_category():
            """选用通用“单位”类别，若缺失则取任一类别兜底。"""
            category = self.env.ref("uom.product_uom_categ_unit", raise_if_not_found=False)
            if not category:
                category = self.env["uom.category"].sudo().search([], limit=1)
            return category

        for row in row_iter:
            
             #小工具：按字段名取这一行对应列的值
            def get(field):
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return ""
                return row[idx] if not isinstance(row, dict) else row.get(idx)

            name = str(get("name") or "").strip()
            code = str(get("code") or "").strip()
            source_code = code or False
            raw_summary_label = next(
                (
                    str(value or "").strip()
                    for value in (row.values() if isinstance(row, dict) else row)
                    if any(
                        key in re.sub(r"\s+", "", str(value or "")).lower()
                        for key in ("合计", "小计", "本页", "本表")
                    )
                ),
                "",
            )
            if not (name or code) and raw_summary_label:
                name = raw_summary_label
            if not (name or code):
                skipped_rows += 1
                continue

            eff_boq_category = boq_category or self.boq_category
            source_label = re.sub(r"\s+", "", f"{code}{name}").lower()
            is_summary = any(key in source_label for key in ("合计", "小计", "本页", "本表"))
            summary_type = "total" if any(key in source_label for key in ("合计", "总计")) else "subtotal"
            is_heading = False
            if is_summary:
                source_name = name or code
                code = f"SUMMARY-{sheet_index or 0:02d}-{len(rows) + 1:04d}"
                name = source_name
            if eff_boq_category == "fee" and "税" in name:
                eff_boq_category = "tax"

            if eff_boq_category in ("fee", "tax") and not code:
                code = f"{eff_boq_category.upper()}-{sheet_index or 0:02d}-{len(rows) + 1:03d}"

            # 逻辑上的“项目编码”：只保留阿拉伯数字，过滤掉序号/符号
            logical_code_digits = re.sub(r"\D", "", code or "")

            # 非分部分项清单：缺少编码直接跳过，避免必填校验失败
            if eff_boq_category != "boq" and not code and not is_summary:
                skipped_rows += 1
                continue

            # 分部分项清单：项目编码空、名称有值（且非数字） -> 记录当前分部，跳过本行
            if (
                eff_boq_category == "boq"
                and not code
                and name
                and not self._is_number(name)
                and not is_summary
            ):
                current_division = name
                is_heading = True
                code = f"HEADING-{sheet_index or 0:02d}-{len(rows) + 1:04d}"

            vals = {
                "project_id": self.project_id.id,
                "sequence": len(rows) + 1,
                "name": name,
                "section_type": self.section_type or section_type or False,
                "single_name": default_single or self.single_name or False,
                "unit_name": default_unit or self.unit_name or False,
                "major_name": major_name or False,
                "sheet_index": sheet_index,
                "sheet_name": sheet_name,
                "boq_category": eff_boq_category or "boq",
                # 默认认为都是清单项，由 _create_with_hierarchy 根据层级修正。
                "line_type": "item",
                "source_row_type": summary_type if is_summary else ("heading" if is_heading else "item"),
                "source_code": source_code if not (is_summary or is_heading) else False,
            }
            if section_type and not vals["section_type"]:
                vals["section_type"] = section_type

            spec = str(get("spec") or "").strip()
            remark = str(get("remark") or "").strip()
            qty = get("quantity")
            price = get("price")
            amount_val = get("amount")
            rate_val = get("rate")
            calc_base = str(get("calc_base") or "").strip()

            # ---- 总价措施 / 规费 / 税金 专用规则 ----
            if eff_boq_category in ("total_measure", "fee", "tax"):
                lower_name = (name or "").lower()
                # 子目计算行也是源文件事实：完整保留，但标记为计算明细，
                # 避免与其上级费用项重复计入清单版本总额。
                if not is_summary and not logical_code_digits and self._is_number(amount_val):
                    code = f"DETAIL-{sheet_index or 0:02d}-{len(rows) + 1:04d}"
                    vals["code"] = code
                    vals["is_calculation_detail"] = True

                # 3) 只有金额，没有工程量/单价 → 用金额补齐 qty/price
                if (not self._is_number(qty)) and (not self._is_number(price)) and self._is_number(amount_val):
                    qty = 1.0
                    price = amount_val

                # 4) 金额型费用行统一视为清单项
                vals["line_type"] = "group" if is_summary else "item"

            if code:
                vals["code"] = code
            if spec:
                vals["spec"] = spec

            # 记录当前分部名称，便于 WBS 直接使用（无需再从 remark 里解析）。
            if current_division:
                vals["division_name"] = current_division
            if remark or current_division:
                prefix = f"[分部]{current_division}" if current_division else ""
                vals["remark"] = f"{prefix} {remark}".strip() if (prefix or remark) else False

            vals["quantity"] = self._to_float(qty)
            vals["price"] = self._to_float(price)
            vals["has_imported_amount"] = self._is_number(amount_val)
            vals["imported_amount"] = self._to_float(amount_val)
            vals["calculated_amount"] = (
                0.0 if is_summary else vals["quantity"] * vals["price"]
            )
            vals["source_calc_base"] = calc_base or False
            vals["has_source_rate"] = self._is_number(rate_val)
            vals["source_rate"] = self._to_float(rate_val)

            # 若数量/单价/合价均为0，则视为标题/小计行跳过
            if strict_numeric and not is_heading:
                if not any(
                    [
                        self._is_number(qty),
                        self._is_number(price),
                        self._is_number(amount_val),
                    ]
                ):
                    skipped_rows += 1
                    continue

            if is_summary or is_heading:
                vals["line_type"] = "group"

            # ===== 计量单位处理 =====
            uom = False
            uom_name = str(get("uom_id") or "").strip()
            if uom_name:
                norm_name = self._normalize_uom_name(uom_name)
                canonical = self._canonical_uom(norm_name)
                search_key = canonical or norm_name or uom_name

                uom = uom_cache.get(search_key)
                create_name = None

                if uom is None:
                    # 先按规范名找
                    uom = Uom.search([("name", "=", search_key)], limit=1)
                    # 再按原始名兜底
                    if not uom and uom_name != search_key:
                        uom = Uom.search([("name", "=", uom_name)], limit=1)

                    if not uom:
                        category = _default_uom_category()
                        if not category:
                            raise UserError(
                                "未找到计量单位类别，无法自动创建单位，请先在系统中创建一个计量单位类别。"
                            )
                        create_name = search_key
                        ref_uom = Uom.search(
                            [
                                ("category_id", "=", category.id),
                                ("uom_type", "=", "reference"),
                            ],
                            limit=1,
                        )
                        uom_vals = {
                            "name": create_name,
                            "category_id": category.id,
                            "factor": 1.0,
                            "factor_inv": 1.0,
                            "rounding": 0.0001,
                            "active": True,
                        }
                        # 如果类别已有参照单位，则新建等效单位用 smaller 并保持 factor=1
                        if ref_uom:
                            uom_vals["uom_type"] = "smaller"
                            uom_vals["factor"] = 1.0
                            uom_vals["factor_inv"] = 1.0
                        else:
                            uom_vals["uom_type"] = "reference"
                        if create_name:
                            created_uoms.add(create_name)
                        if not self.env.context.get("boq_import_preflight"):
                            uom = Uom.create(uom_vals)

                if uom:
                    uom_cache[search_key] = uom

            # 清单文件的总价措施、规费等行经常省略单位。此时必须落到
            # 造价业务语义“项”，不能泄漏 Odoo 基础单位的技术名称 Units。
            if not uom:
                uom = self._business_item_uom()
                if not uom:
                    created_uoms.add("项")

            vals["uom_id"] = uom.id if uom else False

            # ===== 成本项字典匹配 =====
            cost_item_name = str(get("cost_item_id") or "").strip()
            if cost_item_name and Dict:
                cost_item = cost_item_cache.get(cost_item_name)
                if cost_item is None:
                    if isinstance(dict_domain_key, (list, tuple)):
                        domain = list(dict_domain_key)
                    else:
                        domain = [(dict_domain_key, "=", "cost_item")]
                    domain.append(("name", "=", cost_item_name))
                    cost_item = Dict.search(domain, limit=1)
                    cost_item_cache[cost_item_name] = cost_item
                vals["cost_item_id"] = cost_item.id or False

            # --- 总价措施 / 单价措施等“非分部分项”表的分部兜底 ---
            # 这些表本身没有“分部标题行”，为了避免 division_name=False 出现在分组视图里，
            # 这里按清单类别统一给一个可读的分部名称。
            if not vals.get("division_name"):
                if boq_category in ("total_measure",):
                    # 总价措施项目清单
                    vals["division_name"] = "总价措施项目"
                elif boq_category in ("unit_measure",):
                    # 单价措施项目清单
                    vals["division_name"] = "单价措施项目"
                elif boq_category in ("fee", "tax"):
                    vals["division_name"] = "规费及税金"

            rows.append(vals)

        return rows, created_uoms, skipped_rows

    # -------------------------------------------------------------------------
    # 其他项目清单（专用解析）
    # -------------------------------------------------------------------------
    def _build_rows_other(
        self,
        data_rows,
        sheet_index=None,
        sheet_name=None,
        section_type=None,
        default_single=None,
        default_unit=None,
        major_name=None,
    ):
        """
        解析《其他项目清单与计价汇总表》：
        - A 列：序号/层级编码（1 / 2 / 2.1 / 3 / 合计）
        - B 列：项目名称
        - C 列：金额（无数量/单价）
        """
        rows = []
        skipped = 0

        def _default_uom():
            return self._business_item_uom()

        default_uom = _default_uom()

        for row in data_rows:
            code = str((row[0] if len(row) > 0 else "") or "").strip()
            name = str((row[1] if len(row) > 1 else "") or "").strip()
            amount_raw = row[2] if len(row) > 2 else ""

            # 空行直接跳过
            if not code and not name:
                skipped += 1
                continue

            # 合计/总计行（写在序号或项目名称里）一律跳过
            label = f"{code}{name}".replace(" ", "").replace("　", "")
            if label in ("合计", "本表合计", "本页合计", "总计", "台计"):
                skipped += 1
                continue

            line_type, level = self._parse_other_line_level(code)
            if not line_type or not level:
                # 兜底：当作一级标题
                line_type = "group"
                level = 1

            amount = self._to_float(amount_raw)
            has_source_amount = self._is_number(amount_raw)
            if has_source_amount:
                # G.1 中的一级序号可能是“暂列金额”等真实计价项，
                # 有明确来源金额时不得因序号层级被降级为标题。
                line_type = "item"
            qty = 1.0
            price = amount

            vals = {
                "project_id": self.project_id.id,
                "name": name,
                "code": code,
                "quantity": qty,
                "price": price,
                "has_imported_amount": has_source_amount,
                "imported_amount": amount,
                "calculated_amount": amount,
                "boq_category": "other",
                "division_name": "其他项目",
                "line_type": line_type,
                "source_row_type": "heading" if line_type == "group" else "item",
                "source_code": code if line_type == "item" else False,
                "sheet_index": sheet_index,
                "sheet_name": sheet_name,
                "section_type": self.section_type or section_type or False,
                "single_name": default_single or self.single_name or False,
                "unit_name": default_unit or self.unit_name or False,
                "major_name": major_name or False,
            }
            if default_uom:
                vals["uom_id"] = default_uom.id
            rows.append(vals)

        return rows, skipped

    # -------------------------------------------------------------------------
    # 字符串/数值工具
    # -------------------------------------------------------------------------
    def _read_as_csv(self, data_bytes):
        """Return CSV string from raw bytes."""
        return self._parse_csv_bytes(data_bytes)

    def _parse_csv_bytes(self, data_bytes):
        """Try utf-8, then gbk."""
        for encoding in ("utf-8", "gbk"):
            try:
                return data_bytes.decode(encoding)
            except Exception:
                continue
        raise UserError("无法解码导入文件，请确认使用 UTF-8 或 GBK 编码。")

    @staticmethod
    def _guess_section_type(sheet_title):
        title = (sheet_title or "").lower()
        mapping = {
            "build": "building",
            "建筑": "building",
            "机电": "installation",
            "安装": "installation",
            "elect": "installation",
            "装饰": "decoration",
            "decoration": "decoration",
            "景观": "landscape",
            "landscape": "landscape",
        }
        for key, val in mapping.items():
            if key in title:
                return val
        return False

    def _is_supported_sheet(self, title):
        """
        只要能识别出 sheet_type，就认为是“清单相关的有效 sheet”，其他一律跳过
        （封面、汇总表、投标总说明之类不会被读取）。
        """
        sheet_type, category = ProjectBoqImportWizard._classify_sheet_title(title or "")
        return bool(sheet_type)

    @staticmethod
    def _map_major_to_section_type(major_name):
        """根据专业名称映射工程类别（section_type）"""
        text = (major_name or "").lower()
        mapping = {
            "装饰": "decoration",
            "装修": "decoration",
            "建筑": "building",
            "土建": "building",
            "给排水": "installation",
            "暖通": "installation",
            "电气": "installation",
            "强电": "installation",
            "弱电": "installation",
            "机电": "installation",
            "消防": "installation",
            "安装": "installation",
            "景观": "landscape",
            "市政": "municipal",
        }
        for key, val in mapping.items():
            if key in text:
                return val
        return False

    # 其他项目清单：根据序号判断层级与行类型
    @staticmethod
    def _parse_other_line_level(code):
        """
        返回 (line_type, level)：
        - 纯数字：一级 group
        - 数字.数字：二级 item
        - 合计：None
        """
        text = (code or "").strip()
        if text in ("合计", "合 计", "台计"):
            return None, None
        if not text:
            return None, None
        if text.isdigit():
            return "group", 1
        if re.match(r"^\d+\.\d+$", text):
            return "item", 2
        return None, None

    @staticmethod
    def _detect_boq_category(sheet_title):
        """
        根据 sheet 名推断清单类别：分部分项/单价措施/总价措施/规费/税金/其他项目。
        实现上复用 _classify_sheet_title 的逻辑。
        """
        sheet_type, category = ProjectBoqImportWizard._classify_sheet_title(sheet_title or "")
        if category:
            return category
        # 没识别到就按分部分项兜底
        return "boq"

    @staticmethod
    def _normalize_header(title):
        text = str(title or "").strip()
        text = re.sub(r"\s+", "", text)
        return text.lower()

    @staticmethod
    def _is_number(val):
        try:
            if isinstance(val, str):
                cleaned = ProjectBoqImportWizard._clean_number_str(val)
                if cleaned in ("", "-", "--"):
                    return False
                float(cleaned)
            else:
                float(val)
            return True
        except Exception:
            return False

    @staticmethod
    def _to_float(val):
        try:
            if isinstance(val, str):
                cleaned = ProjectBoqImportWizard._clean_number_str(val)
                return float(cleaned or 0.0)
            return float(val or 0.0)
        except Exception:
            return 0.0

    def _get_dictionary_model(self):
        """返回可用的字典模型及类型字段键。"""
        dict_model = "project.dictionary" if "project.dictionary" in self.env.registry else "sc.dictionary"
        if dict_model not in self.env.registry:
            return None, None
        Dict = self.env[dict_model]
        fields_map = Dict._fields
        if "type" in fields_map:
            return Dict, "type"
        if "type_id" in fields_map:
            return Dict, "type_id.code"
        return Dict, "type"

    @staticmethod
    def _clean_number_str(text):
        """Remove common thousand separators and spaces."""
        cleaned = str(text or "")
        cleaned = cleaned.replace(",", "").replace("，", "").replace(" ", "").strip()
        return cleaned

    # --- UoM helpers ---
    @staticmethod
    def _normalize_xls_diagnostics(raw):
        """Deduplicate xlrd/OLE2 diagnostics for preflight evidence."""
        diagnostics = []
        seen = set()
        for line in str(raw or "").splitlines():
            normalized = re.sub(r"\s+", " ", line).strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            diagnostics.append(normalized)
        return diagnostics

    def _business_item_uom(self):
        """Return/create the BOQ lump-sum unit without exposing Odoo's Units label."""
        category = self.env.ref("uom.product_uom_categ_unit", raise_if_not_found=False)
        if not category:
            category = self.env["uom.category"].search([], limit=1)
        if not category:
            return self.env["uom.uom"]

        Uom = self.env["uom.uom"]
        item_uom = Uom.search(
            [("category_id", "=", category.id), ("name", "=", "项")], limit=1
        )
        if item_uom or self.env.context.get("boq_import_preflight"):
            return item_uom
        ref_uom = Uom.search(
            [("category_id", "=", category.id), ("uom_type", "=", "reference")], limit=1
        )
        return Uom.create(
            {
                "name": "项",
                "category_id": category.id,
                "uom_type": "smaller" if ref_uom else "reference",
                "factor": 1.0,
                "factor_inv": 1.0,
                "rounding": 0.0001,
                "active": True,
            }
        )

    def _normalize_uom_name(self, name):
        """基本规范化：去空格、全角转半角、小写。"""
        text = misc.ustr(name or "").strip()
        text = re.sub(r"\s+", "", text)
        res = []
        for ch in text:
            code = ord(ch)
            if 0xFF01 <= code <= 0xFF5E:
                code -= 0xfee0
                ch = chr(code)
            res.append(ch)
        return "".join(res).lower()

    def _canonical_uom(self, norm_name):
        """根据别名映射返回规范名，否则返回原名。"""
        for main, aliases in self.UOM_ALIAS_MAP.items():
            if norm_name == main:
                return main
            for alias in aliases:
                if norm_name == self._normalize_uom_name(alias):
                    return main
        return norm_name

    # -------------------------------------------------------------------------
    # 批量创建 & 层级构建
    # -------------------------------------------------------------------------
    def _batch_create(self, model, vals_list):
        """批量创建，避免一次性巨大列表占用内存/锁时间过长。"""
        if not vals_list:
            return 0
        size = self.BATCH_CREATE_SIZE or 500
        created = 0
        for start in range(0, len(vals_list), size):
            chunk = vals_list[start : start + size]
            model.create(chunk)
            created += len(chunk)
        return created

    @staticmethod
    def _summary_calculation_values(scope):
        rows = list(scope)
        return {
            "calculated_amount": sum(
                (row.quantity or 0.0) * (row.price or 0.0) for row in rows
            ),
            "calculation_scope_item_count": len(rows),
            "calculation_scope_start_sequence": rows[0].sequence if rows else 0,
            "calculation_scope_end_sequence": rows[-1].sequence if rows else 0,
        }

    def _finalize_source_summary_calculations(self, lines):
        """Freeze an auditable source-vs-system comparison for each summary row."""
        by_sheet = {}
        for line in lines.sorted(lambda row: (row.sheet_index or 0, row.sequence or 0, row.id)):
            by_sheet.setdefault(line.sheet_index or 0, []).append(line)
        for sheet_lines in by_sheet.values():
            all_items = []
            subtotal_window = []
            for line in sheet_lines:
                if line.line_type == "item":
                    all_items.append(line)
                    subtotal_window.append(line)
                    if not line.calculated_amount:
                        line.calculated_amount = (line.quantity or 0.0) * (line.price or 0.0)
                    continue
                if line.source_row_type not in ("subtotal", "total"):
                    continue
                scope = all_items if line.source_row_type == "total" else subtotal_window
                line.write(self._summary_calculation_values(scope))
                if line.source_row_type == "subtotal":
                    subtotal_window = []

    def _create_with_hierarchy(self, model, vals_list):
        """
        批量创建 + 根据编码/上下文推断层级，写入 parent_id + line_type。
        仅在分部分项清单（boq_category='boq'）中使用。

        思路：
        - 保持创建顺序（vals_list 的顺序），避免打乱 Excel 原始行序；
        - 以 (project, section_type, single_name, unit_name, sheet_index) 为分段 key，
          每一段单独维护一个“层级栈”（stack: level -> record）；
        - 根据编码/名称/是否有数量，调用 _classify_line 得到 (line_type, level)；
        - level=0 无 parent，level>0 时 parent = stack[level-1]；
        - 最后将当前记录放入 stack[level]，供后面的行作为下级挂接。
        """
        if not vals_list:
            return 0

        # 一次性创建所有记录（保持 vals_list 顺序）
        records = model.create(vals_list)

        # 为了保证层级构建稳定，以 sheet_index / id 排个序
        ordered_records = sorted(
            records,
            key=lambda r: (
                r.project_id.id or 0,
                r.section_type or "",
                r.single_name or "",
                r.unit_name or "",
                r.sheet_index or 0,
                r.id,
            ),
        )

        current_key = None
        stack = {}  # level(int) -> record

        for rec in ordered_records:
            key = (
                rec.project_id.id,
                rec.section_type or "",
                rec.single_name or "",
                rec.unit_name or "",
                rec.sheet_index or 0,
            )
            # 换了 sheet / 单项工程 / 单位工程：重置层级栈
            if key != current_key:
                stack = {}
                current_key = key

            # 其他项目清单：使用专用层级规则（code 决定 level）
            if rec.boq_category == "other":
                o_line_type, o_level = ProjectBoqImportWizard._parse_other_line_level(rec.code)
                if not o_line_type or o_level is None:
                    o_line_type, o_level = "group", 1
                if rec.source_row_type == "item":
                    o_line_type = "item"
                rec.line_type = o_line_type
                if o_level <= 0:
                    rec.parent_id = False
                else:
                    parent = stack.get(o_level - 1)
                    rec.parent_id = parent.id if parent else False
                stack[o_level] = rec
                continue

            if rec.source_row_type in ("heading", "subtotal", "total"):
                rec.line_type = "group"
                rec.parent_id = False
                continue

            line_type, level = self._classify_line(
                code=rec.code,
                name=rec.name,
                qty=rec.quantity,
                price=rec.price,
                amount=rec.amount,
                boq_category=rec.boq_category,
            )

            # 行的业务类型以解析契约为准；编码长度只用于定位层级。
            # 否则 6 位补充项编码会被误判为分组并从金额汇总中丢失。
            rec.line_type = "item" if rec.source_row_type == "item" else line_type

            # 处理 parent_id
            if level <= 0:
                rec.parent_id = False
            else:
                parent = stack.get(level - 1)
                rec.parent_id = parent.id if parent else False

            # 记录当前层级最近一行，供后面的子级挂接。
            stack[level] = rec

        return len(records)

    @staticmethod
    def _classify_line(code, name, qty, price, amount, boq_category):
        """
        读一行，判断：
        - line_type: major / division / group / item
        - level: 0,1,2,3  对应 章 / 分部 / 小结 / 清单项

        规则分两块：
        A) 总价措施 / 规费 / 税金 等“金额型费用表”
        B) 普通分部分项清单（编码驱动层级）
        """
        code = (code or "").strip()
        name = (name or "").strip()
        lname = name.lower()

        # 是否有“数量/单价/金额”数值
        has_numeric = any(
            ProjectBoqImportWizard._is_number(v)
            for v in (qty, price, amount)
        )

        # ---------- A) 总价措施 / 规费 / 税金 ----------
        if boq_category in ("total_measure", "fee", "tax"):
            # 典型结构：第一行有项目编码，下面若干行无编码，表示该费用的组成明细
            #   1  安全文明施工费  code=0411..., amount 有值
            #   1.1 环境保护费    code 为空，费率/金额有值
            #   1.2 文明施工费    ...
            #
            # 策略：
            # - 只要这一行有 code → 视为“费用汇总行”（group, level 1）
            # - 同一表中，随后无 code 的行 → 视为该费用下的明细项（item, level 2）
            # - “合计/本表合计”等仍按通用小计过滤逻辑跳过（在别处已经处理）
            if code:
                return "group", 1
            else:
                # 无编码，但有金额/费率 → 明细项
                if has_numeric:
                    return "item", 2
                # 实在啥都没有，就当成标题行（几乎不会出现）
                return "group", 1

        # ---------- B) 普通分部分项清单：原有逻辑 ----------
        # 特殊前缀：MAJ-xxx / DIV-xxx
        if code.startswith("MAJ-") or code.startswith("MAJ"):
            return "major", 0
        if code.startswith("DIV-") or code.startswith("DIV"):
            return "division", 1

        # 纯数字/带点编码
        pure = code.replace(".", "")
        if pure.isdigit():
            length = len(pure)
            if length <= 2:
                return "major", 0
            if length <= 4:
                return "division", 1
            if length <= 6:
                return "group", 2
            # 一般 8~12 位都是具体清单项目码
            return "item", 3

        # 无编码 & 无数值：标题/汇总
        if not code and not has_numeric:
            # 合计/小计 视为小结行
            if any(k in lname for k in ("合计", "小计", "本页", "本表")):
                return "group", 2
            # 名称里包含“工程/专业”等，视为分部工程
            if any(k in name for k in ("工程", "专业", "道路", "市政", "桥梁", "绿化")):
                return "division", 1
            # 其他标题，暂当 group
            return "group", 2

        # 有数值，但是编码非数字（比如 “0401090024-01” 的特殊写法）
        if has_numeric:
            return "item", 3

        # 兜底：实在分不清的都当清单项，不破坏数据
        return "item", 3


# -------------------------------------------------------------------------
# 层级构建器（封装栈操作，保持现有层级算法）
# -------------------------------------------------------------------------
class HierarchyBuilder:
    def __init__(self):
        self.stack = {}

    def reset(self):
        self.stack = {}

    def place(self, rec, level):
        parent = self.stack.get(level - 1)
        rec.parent_id = parent.id if parent else False
        self.stack[level] = rec

    def refresh_parent_path(self, records):
        """统一刷新 parent_path；失败时保持导入流程不中断。"""
        try:
            records._parent_store_compute()
        except Exception:
            _logger.debug("Unable to refresh BOQ hierarchy parent path.", exc_info=True)

    def heal_hierarchy(self, records):
        """
        层级连续性修复扩展点。
        当前保持导入行为稳定，可用于统一调整 level/parent_id/display_order。
        """
        return records


# -------------------------------------------------------------------------
# 导入解析适配层（行为保持不变）
# -------------------------------------------------------------------------
class RowParser:
    """行解析适配器；可按清单类别扩展并保持默认行为稳定。"""

    def __init__(self, wizard):
        self.wizard = wizard

    def parse_row(self, raw_row, col_map):
        """返回原始行，供类别化解析扩展使用。"""
        return raw_row


class BoqParser:
    """
    导入解析适配层。
    当前仍委托原有 _parse_excel/_build_rows_from_iter，承担结构封装与章节池收集，
    不改变导入业务行为。
    """

    def __init__(self, wizard):
        self.wizard = wizard
        self.row_parser = RowParser(wizard)
        # 章节池：收集标题/章节文本，仅收集候选，不做层级推断。
        self.chapter_pool = []

    def parse_file(self, data, filename, include_details=False):
        """按文件类型分发，返回 rows/created_uoms/skipped。"""
        fname = (filename or "").lower()
        if fname.endswith((".xlsx", ".xls")):
            return self.wizard._parse_excel(data, fname, include_details=include_details)
        # CSV 默认解析
        content = self.wizard._read_as_csv(data)
        result = self.wizard._parse_csv_content(content)
        if include_details:
            return (
                *result,
                {
                    "schema": "sc.boq.full_import.v1",
                    "analyses": [],
                    "summary_components": [],
                },
            )
        return result

    def parse_sheet(self, sheet, sheet_index):
        """
        保留工作表解析扩展点，当前由 wizard._parse_excel 处理。
        预解析合并单元格标题区并收集章节池（仅收集，不推断）。
        """
        titles = self.parse_merged_title_area(sheet)
        if titles:
            self.chapter_pool.extend(titles)
        return None

    def parse_rows(self, data_rows):
        """保留行解析扩展点，当前由 wizard._build_rows_from_iter 处理。"""
        return None

    # ------------------ 章节/标题预解析 ------------------
    def parse_merged_title_area(self, sheet, max_rows=5):
        """
        简单读取前几行合并单元格的非空文本，作为章节池候选。
        仅收集文本，不做层级推断。
        """
        titles = []
        try:
            merge_ranges = sheet.merged_cells
        except Exception:
            merge_ranges = None
        ranges = getattr(merge_ranges, "ranges", merge_ranges) or []
        seen = set()
        for m in ranges:
            try:
                min_row, min_col, max_row, max_col = m.min_row, m.min_col, m.max_row, m.max_col
            except Exception:
                continue
            if min_row > max_rows:
                continue
            try:
                val = sheet.cell(row=min_row, column=min_col).value
            except Exception:
                val = None
            text = str(val or "").strip()
            if text and text not in seen:
                titles.append(text)
                seen.add(text)
        # 兜底：再扫一遍前 max_rows 行的非空单元格，补充章节文本
        for r in range(1, max_rows + 1):
            row_vals = []
            try:
                row_vals = [str(sheet.cell(row=r, column=c).value or "").strip() for c in range(1, (sheet.max_column or 0) + 1)]
            except Exception:
                _logger.debug("Unable to scan BOQ merged title row.", exc_info=True)
            for v in row_vals:
                if v and v not in seen:
                    titles.append(v)
                    seen.add(v)
        return titles
