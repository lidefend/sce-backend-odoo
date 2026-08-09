"""Small, read-only workbook facade for native norm imports.

The parser consumes one 1-based cell API regardless of whether the uploaded
file is an OOXML workbook or a legacy BIFF/OLE2 workbook.  Business parsing
therefore stays format-neutral and is shared by preview and import.
"""

import io

import xlrd
from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = (".xls", ".xlsx", ".xlsm")


class _XlsCell:
    def __init__(self, value):
        self.value = value


class _XlsSheet:
    def __init__(self, sheet):
        self._sheet = sheet
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, *, row, column):
        if row <= 0 or column <= 0 or row > self.max_row or column > self.max_column:
            return _XlsCell(None)
        return _XlsCell(self._sheet.cell_value(row - 1, column - 1))


class _XlsWorkbook:
    def __init__(self, workbook):
        self._workbook = workbook
        self.sheetnames = list(workbook.sheet_names())

    def __getitem__(self, sheet_name):
        return _XlsSheet(self._workbook.sheet_by_name(sheet_name))

    def close(self):
        self._workbook.release_resources()


def open_workbook(data, filename):
    """Return an openpyxl-compatible, read-only workbook surface."""
    lower_name = str(filename or "").strip().lower()
    if lower_name.endswith(".xls"):
        return _XlsWorkbook(xlrd.open_workbook(file_contents=data, on_demand=True))
    return load_workbook(io.BytesIO(data), data_only=True, read_only=True)
